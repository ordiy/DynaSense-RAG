"""
Extract plain text from XLSX bytes for ingestion.

Design mirrors pdf_extract.py: keep parsing isolated from FastAPI and rag_core so
the pipeline always receives UTF-8 text — the same path as PDF/TXT/DOCX/MD for
chunking, embeddings, and graph ingest.

Extraction strategy:
  - Each worksheet becomes a section headed by "## Sheet: <name>"
  - Rows are emitted as tab-separated values (header row first if present)
  - Empty rows and fully-empty sheets are skipped
  - Multiple sheets are joined with double newlines

This format is intentionally human-readable so that the Jina segmenter can
produce meaningful semantic chunks and the LLM triple extractor can identify
entities/relationships from structured tabular data.
"""
from __future__ import annotations

from io import BytesIO


class XlsxExtractError(ValueError):
    """Raised when the file is not a valid XLSX or cannot be read."""


def extract_text_from_xlsx_bytes(data: bytes) -> str:
    """
    Extract text from an XLSX file given its raw bytes.

    Returns a single string with each sheet as a named section and rows as
    tab-separated lines. Raises XlsxExtractError for invalid/empty files.
    """
    if not data:
        raise XlsxExtractError("Empty file.")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise XlsxExtractError(f"Invalid or corrupted XLSX: {e}") from e

    sheet_texts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        row_texts: list[str] = []

        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            # Skip entirely empty rows
            if not any(c.strip() for c in cells):
                continue
            row_texts.append("\t".join(cells))

        if row_texts:
            section = f"## Sheet: {sheet_name}\n" + "\n".join(row_texts)
            sheet_texts.append(section)

    wb.close()

    if not sheet_texts:
        return ""

    return "\n\n".join(sheet_texts).strip()
